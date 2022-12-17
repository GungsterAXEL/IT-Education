'''
Написать функцию для добавления новой записи в книгу, и сохранения в базу данных.
На входе минимум 4 переменных: Имя, Фамилия, номер и примечание.
'''
import os
import Logger
import openpyxl
os.chdir(os.path.dirname(__file__))


phonebook = openpyxl.open('PhoneBook.xlsx', read_only=True)
sheet = phonebook.active
sheet_2 = phonebook.worksheets[2]
print(sheet_2['A2'])

#for row in range(1,sheet.max_row + 1):
#	name = sheet[row][0].value
#	sirname = sheet[row][1].value
#	number = sheet[row][2].value
#	footnote = sheet[row][3].value
#	print(row, name, sirname, number, footnote)

#cells = sheet['A1':'C5']
#for name, sirname, number in cells:
#	print(name.value, sirname.value, number.value)

#for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row + 1, min_col=1, max_col=4):
#	print(row)
#	for cell in row:
#		print(cell.value, end= ' ')
#	print()

def new_entry_saver(first_name, last_name, tel_number, footnote):
	new_entry = f'{first_name}, {last_name}, {tel_number}, {footnote};\n'
	try:
		with open('PhoneBook.txt', 'a', encoding='utf-8') as phone_data:
			phone_data.write(new_entry)
			Logger.log_logger('New_Entry_Saver', True)
	except:
		Logger.log_logger('New_Entry_Saver', False)